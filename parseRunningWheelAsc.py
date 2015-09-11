#! /usr/bin/env python

# Python script for parsing running wheel data output

import re # import regular expession library
import sys # import OS/system level tools
import csv # import CSV tools
from datetime import datetime, timedelta
from openpyxl import Workbook # import python Excel parser
from openpyxl import load_workbook # for loading excel files

# If no arguments are entered...
USAGE = '''\n***parseRunningWheelAsc.py script - by Prech Brian Uapinyoying***\n
Parses mouse running wheel data that has been output into a ASCII file
directly obtained from the Vitalview Software.  The input file should have two 
headers. The first will be the the over all file header that says 'Experiment 
Logfile:' followed by the date. Second, the data header below in comma 
delimited format (csv).

The script will output a csv file for each of the following:
1) Mice data - summary statistics
2) Raw data - unmanipulated
3) Distance calculated - converted turns to meters
4) Time Filtered - includes data > 18:01:00 of day 1 to 06:00:00 of day 3
5) Sum Hourly - minute data is summed into hours per row
6) Cumulative - data from each hour is compounded onto the previous

Finally, a single Excel file containing all the output csv files is created.

Command Usage:
> parseRunningWheelAsc.py <input.csv>\n'''

# Regular Expressions and static (unchanging) variables 
SAMPLE_NAME_REGEXP = r'([\w \/]+) Turns Data'
FILE_NAME_REGEXP = r'(.+)\..+'
HEADER_STRING = 'Experiment Logfile:'

### Here lie the functions or reusable chunks of code.
### it also helps keep the code a bit cleaner in the main body below

def get_file_name(FILE_NAME_REGEXP):
    """Function to get the whole file name and the filename without the .asc"""
    wholeFileName = sys.argv[1] # get filename from first command line argument
    
    # Use a regular expression to match the filename
    nameSearchObj = re.search(FILE_NAME_REGEXP, wholeFileName)

    # Capture part without extension. Notice FILE_NAME_REGEXP in parentheses
    # is inside group(1). If there was another pair of () it would be group(2)
    fileNameNoExtension = nameSearchObj.group(1)
    return wholeFileName, fileNameNoExtension # return information in a pair


def check_fileHeader(readerObj, HEADER_STRING):
    """Function to check the file's header."""
    #Move to the next (first) row of data and assign it to a variable
    firstHeader = readerObj.next()

    # Try a search and find on the header row to match HEADER_STRING. If return
    # error, print feedback and quit program.
    try:
        searchHeader = re.search(HEADER_STRING, firstHeader[0])
        searchResult = searchHeader.group(0) # group(0) returns whole string
        return firstHeader 
    except AttributeError:
        print "ERROR: This file is in the wrong format or is the wrong file type."
        sys.exit(0) # quit the program


def check_header2(SAMPLE_NAME_REGEXP, rowOfData):
    """Function to check the second header and formats header for distance csv
    Returns raw header and reformatted header"""
    rowLength = len(rowOfData)

    # Use the modulo (%) operator which yeilds remainder after dividing by a
    # number. Sample data should be in triplicate columns. If it's not, quit.
    # Else, reformat the header and return formated one
    if rowLength % 3 != 0:
        print "ERROR: Sample data are not in triplicate columns."
        sys.exit(0) # quit the program
    else:
        # For distance file, reformat header
        distHeaderRow = ["Date", "Time"]

        # Loop through the columns of data starting at col 3, until end. Step 3
        # at a time to get col containing Turn Data for each sample only.
        # Remember python lists start at 0, 1, 2, ...
        # I reformat the column order here to remove duplicate time and dates
        # and add a meters per minute one.
        for i in xrange(2, rowLength, 3):
            sampleHeader = rowOfData[i]

            # Add the Turns Data sample header as is to the row
            distHeaderRow.append(sampleHeader)

            # Capture the sample name (without 'Turn Data')
            searchObj = re.search(SAMPLE_NAME_REGEXP, sampleHeader)
            searchResult = searchObj.group(1)

            # Append meters/min to end of name
            sampleDistHeader = searchResult + ' meters/min'

            # Add this new meters/min sample header as a new column
            distHeaderRow.append(sampleDistHeader)
        # Return True for header checked, and the newly created header
        return distHeaderRow


def fill_transpose(listOfStreaks):
    """ Takes a list of lists (sample data arranged longitudinally), each sublist 
    starts with a header. Converts it to column format (vertical arrangement)
    to make it easy to print"""

    # Find length of longest sublist
    listLen = len(listOfStreaks)
    maxSubLen = 0
    for l in listOfStreaks:
        subListLen = len(l)
        if subListLen > maxSubLen:
            maxSubLen = subListLen

    # Need to even out the row lengths first, make new list of lists that have
    # even number of rows per column
    matrixOfStreaks = []
    for m in range(0, listLen):
        matrixOfStreaks.append([])
        for n in range(0, maxSubLen):
            matrixOfStreaks[m].append('')

    #print matrixOfStreaks

    # # Fill new matrix with original data
    for p in range(0, listLen):
        for q in range(0, len(listOfStreaks[p])):
            matrixOfStreaks[p][q] = listOfStreaks[p][q]


    # Generate a matrix wilth all equal rows, fill blanks with emptys ''
    transposedList = []
    for i in range(0, maxSubLen):
        transposedList.append([])
        for j in range(0, listLen):
            transposedList[i].append('')

    # Now transfer data from listOfStreaks to new transposedList in right order
    for x in range(0, listLen):
        for y in range(0, len(listOfStreaks[x])):
            transposedList[y][x] = listOfStreaks[x][y]
    #print transposedList

    return transposedList

# Track number of arguments given with command
numArgs = len(sys.argv)

### Main body of program/script ###

# Check number arguments given is 2 (program + file), else print USAGE
if numArgs == 2:
    
    # Set a couple of variables for keeping track of rows and columns
    miceRowNum = 0 # keep track of row numbers of mice data file

    # Grab the entire file name and the name without the (.asc) extension
    wholeFileName, fileNameNoExtension = get_file_name(FILE_NAME_REGEXP)
    # Make a few file names
    miceCsvName = fileNameNoExtension + '_mice.csv'
    rawCsvName = fileNameNoExtension + '_rawData.csv'
    distCsvName = fileNameNoExtension + '_distData.csv'

    # Open a connection to the provided csv file to read from
    with open(wholeFileName, 'rb') as csvFile:
        # Turns the open file into an object we can use to pull data from
        originalFileReader = csv.reader(csvFile, delimiter=",", quotechar='"')

        # Check first file header and assign to variable
        fileHeader = check_fileHeader(originalFileReader, HEADER_STRING)

        # Create a mice data csv file
        with open(miceCsvName, 'wb') as miceOutFile:
            miceFileWriter = csv.writer(miceOutFile)

            # Create a raw data csv file
            with open(rawCsvName, 'wb') as rawOutFile:
                rawFileWriter = csv.writer(rawOutFile)

                # Create a calculated distance (meters/min) file
                with open(distCsvName, 'wb') as distOutFile:
                    distFileWriter = csv.writer(distOutFile)

                    # First time around?
                    firstRowOfFile = True

                    # Make sure to check the data header too
                    checkedHeader2 = False

                    # Iterate through every line (row) of the original file
                    for row in originalFileReader:
                        # Put all mouse summary data into the mice sheet, data
                        # should be in less than 3 columns
                        if len(row) < 3:
                            # If very first row, place the checked header
                            if firstRowOfFile:
                                miceFileWriter.writerow(fileHeader)
                                # Next round, set to false to skip this part
                                firstRowOfFile = False
                            # For the rest of the rows, just dump
                            miceFileWriter.writerow(row)
                    
                        # Real data should have at least 3 columns
                        else:
                            # Check second header, is it a multiple of 3?
                            if not checkedHeader2:
                                distHeader = check_header2(
                                    SAMPLE_NAME_REGEXP, row)

                                checkedHeader2 = True
                                
                                # Raw data file takes header row as is
                                rawFileWriter.writerow(row)
                                       
                                # Distance data file takes the formatted one
                                distFileWriter.writerow(distHeader)

                                print "File in correct format."

                            # Once the header is good, parse the data and
                            # calculate the meters/min
                            else:
                                rawFileWriter.writerow(row)
                                distanceRow = [row[0], row[1]]
                                for i in xrange(2, len(row), 3):
                                    sampleData = row[i]
                                    distanceRow.append(sampleData)
                                    meterData = float(sampleData) * 0.361
                                    distanceRow.append(meterData)
                                distFileWriter.writerow(distanceRow)
    
    ### Part 2: Filter the distance file by the time of day, sum up hours and
    ### turn them into cumulative data

    filterCsvName = fileNameNoExtension + '_filterData.csv'
    hourlyCsvName = fileNameNoExtension + '_hourlyData.csv'
    cumulativeCsvName = fileNameNoExtension + '_cumulativeData.csv'
    runStreaksCsvName = fileNameNoExtension + '_runStreaksData.csv'

    # Reopen distance csv file, this time read from it
    with open(distCsvName, 'rb') as distCsvFile:
        # Turns the open file into an object we can use to pull data from
        distFileReader = csv.reader(distCsvFile, delimiter=",", quotechar='"')

        with open(filterCsvName, 'wb') as filterOutFile:
            filterFileWriter = csv.writer(filterOutFile)

            with open(hourlyCsvName, 'wb') as hourlyOutFile:
                hourlyFileWriter = csv.writer(hourlyOutFile)

                with open(cumulativeCsvName, 'wb') as cumulativeOutFile:
                    cumulativeFileWriter = csv.writer(cumulativeOutFile)

                    with open(runStreaksCsvName, 'wb') as runStreaksOutFile:
                        runStreaksFileWriter = csv.writer(runStreaksOutFile)

                        # Grab the deader and then the first row of data
                        header = distFileReader.next()
                        distRow1 = distFileReader.next()

                        # Write the header directly to the files
                        filterFileWriter.writerow(header)
                        hourlyFileWriter.writerow(header)
                        cumulativeFileWriter.writerow(header)

                        # Instantiate some empty list variables for use downstream
                        # Used to help keep track of lists (rows) of numbers
                        hourlyRow = []
                        currListString = []
                        currList = []
                        prevSumList = []
                        tempList = []
                        lastDateTimeList = []

                        runningSumList = []
                        sumTempList = []

                        #### For maxStreak and maxSpeed
                        currVal = [] # temp vals
                        currStreak = [] # temp vals
                        currTimeOff = []
                        currMaxList = []
                        prevMaxList = []
                        maxVal = []  # the row we want
                        maxStreak = []
                        maxTimeOffWheel = []

                        # Determine number of columns minus the date and time fields
                        distRowDataLen = len(distRow1) - 2
                        
                        # Start a list of lists for calculating average running
                        # streaks
                        listOfStreaks = []

                        # Populate the list variables with 0's for eac number of
                        # columns
                        t = 0
                        while t < distRowDataLen:
                            currVal.append(0.0)
                            maxVal.append(0.0)
                            currStreak.append(0.0)
                            maxStreak.append(0.0)
                            maxTimeOffWheel.append(0.0)
                            currTimeOff.append(0.0)
                            prevMaxList.append(0.0)
                            listOfStreaks.append([])
                            t += 1

                        # Set our filter criteria, start after 6pm day 1
                        # end at 6am day 3
                        startDateTimeString = '%s %s' % (distRow1[0], '18:00:00')
                        startDateTime = datetime.strptime(startDateTimeString, '%m/%d/%y %H:%M:%S')
                        endDateTime = startDateTime + timedelta(days=3) - timedelta(hours=12)

                        hourNum = 1
                        distRowNum = 1
                        filterRowNum = 1
                        # loop through all distRows in the distance csv our source of data
                        for distRow in distFileReader:
                            
                            # CFuse date and time information
                            # Now parse the date and time into a datetime object
                            currdistRowDateTimeString = '%s %s' % (distRow[0], distRow[1])
                            currdistRowDateTime = datetime.strptime(currdistRowDateTimeString, '%m/%d/%y %H:%M:%S')

                            # If the date and time are within our criteria...
                            
                            if currdistRowDateTime >= startDateTime and currdistRowDateTime <= endDateTime:
                                # For filter file, write filtered distRows directly to file
                                filterFileWriter.writerow(distRow)

                                # Now to get hourly data, needs some formatting
                                # Grab only turn data and meter columns 
                                # (skip first 2 columns containing date and time)
                                currListString = distRow[2:]
                                
                                # Convert the text numbers to floating point nums
                                for j in currListString:
                                    temp = float(j)
                                    currList.append(temp)

                                # Save this value for later average streak calculations
                                # if filterRowNum == 1:
                                #     prevMaxList = currList
                                if filterRowNum > 1:
                                    prevMaxList = currMaxList   
                                currMaxList = currList
                                


                                # If less than 60 minutes
                                if distRowNum < 60:
                                    # If first minute of hour put current list of numbers
                                    # into the previous sum directly
                                    if distRowNum == 1:
                                        prevSumList = currList
                                        #print distRowNum, currList
                                        currList = []
                                    # # Otherwise, add current numbers to previous numbers
                                    else:
                                        for x, y in zip(prevSumList, currList):
                                            tempList.append(x+y)
                                        prevSumList = tempList
                                        #print distRowNum, currList
                                        tempList = []  # clear the variables 
                                        currList = []
                                    distRowNum += 1
                                        

                                # If it is the last minute of the hour, sum up
                                # everything, format the row, clear all variables,
                                # and restart the numbering back to 1
                                elif distRowNum == 60:
                                    lastDateTimeList = [distRow[0], distRow[1]]
                                    #print distRowNum, currList
                                    for x, y in zip(prevSumList, currList):
                                        tempList.append(x+y)
                                    prevSumList = tempList

                                    # Sum every 60 mins into an hour
                                    if hourNum == 1:
                                        runningSumList = prevSumList
                                    else:
                                        for x, y in zip(prevSumList, runningSumList):
                                            sumTempList.append(x+y)
                                        runningSumList = sumTempList

                                    hourlyRow = lastDateTimeList + prevSumList
                                    cumulativeRow = lastDateTimeList + runningSumList
                                    
                                    # Reset all variables
                                    tempList = []
                                    currList = []
                                    prevSumList = []
                                    sumTempList = []
                                    lastDateTimeList = []
                                    distRowNum = 1
                                    hourNum += 1  # Except the hours

                                    #print hourlyRow
                                    #sys.exit(0)
                                    # Write the calculated data to the csv files
                                    hourlyFileWriter.writerow(hourlyRow)
                                    cumulativeFileWriter.writerow(cumulativeRow)

                                # get max value and longest running streak and rest
                                # streak for each sample
                                # Loop through each distRow and compare values
                                
                                for i in range(0, len(currMaxList)): 
                                    #print filterRowNum, currMaxList[1], prevMaxList[1]
                                    
                                    ### to get the max value of each column
                                    if currMaxList[i] > maxVal[i]:
                                        maxVal[i] = currMaxList[i]
                                    
                                    ### This gives you the longest streak
                                    if currMaxList[i] > 0:
                                        currStreak[i] += 1
                                        currTimeOff[i] = 0
                                    if currStreak[i] > maxStreak[i]:
                                        maxStreak[i] = currStreak[i]
                                    if currTimeOff[i] > maxTimeOffWheel[i]:
                                        maxTimeOffWheel[i] = currTimeOff[i]
                                    if currMaxList[i] == 0:
                                        if prevMaxList[i] != 0:
                                            listOfStreaks[i].append(currStreak[i])
                                        currStreak[i] = 0
                                        currTimeOff[i] += 1

                                filterRowNum += 1

                        maxValTitle = ['', 'Max Value']
                        maxValRow = maxValTitle + maxVal 
                        filterFileWriter.writerow('')
                        filterFileWriter.writerow(maxValRow)
                        maxStreakTitle = ['Max Running', 'Streak (mins)']
                        maxStreakRow = maxStreakTitle + maxStreak
                        filterFileWriter.writerow(maxStreakRow)
                        maxTimeOffTitle = ['Max Rest Time', 'Streak (mins)']
                        maxTimeOffRow = maxTimeOffTitle + maxTimeOffWheel
                        filterFileWriter.writerow(maxTimeOffRow)

                        for x in range(0, len(listOfStreaks)):
                            listOfStreaks[x].insert(0, header[x+2])

                        # use the function to transpose data for export
                        allStreakRow = fill_transpose(listOfStreaks)
                        
                        for z in allStreakRow:
                            runStreaksFileWriter.writerow(z)


    print "Finished calculations and generated csv file output."
    print "Dumping all data into excel file. Takes a minute..."

    # Create an Excel workbook to house this stuff
    wb = Workbook()

    # Grab first sheet
    miceSheet = wb.active
    miceSheet.title = "Mice Summary"

    # Create a new sheet for the rest
    rawSheet = wb.create_sheet(title='Raw Data')
    distSheet = wb.create_sheet(title="Calculated Distances")
    filterSheet = wb.create_sheet(title="Time Filtered")
    hourlySheet = wb.create_sheet(title="Summed Hourly")
    cumulativeSheet = wb.create_sheet(title="Cumulative")
    runStreakSheet = wb.create_sheet(title="Running Streaks")

    # Open each file and place it into their excel sheets
    with open(miceCsvName, 'rb') as miceCsvFile:
        miceFileReader = csv.reader(miceCsvFile, delimiter=",", quotechar='"')
        for row in miceFileReader:
            miceSheet.append(row)

    with open(rawCsvName, 'rb') as rawCsvFile:
        rawFileReader = csv.reader(rawCsvFile, delimiter=",", quotechar='"')
        for row in rawFileReader:
            rawSheet.append(row)

    with open(distCsvName, 'rb') as distCsvFile:
        distFileReader = csv.reader(distCsvFile, delimiter=",", quotechar='"')
        for row in distFileReader:
            distSheet.append(row)

    with open(filterCsvName, 'rb') as filterCsvFile:
        filterFileReader = csv.reader(filterCsvFile, delimiter=",", quotechar='"')
        for row in filterFileReader:
            filterSheet.append(row)

    with open(hourlyCsvName, 'rb') as hourlyCsvFile:
        hourlyFileReader = csv.reader(hourlyCsvFile, delimiter=",", quotechar='"')
        for row in hourlyFileReader:
            hourlySheet.append(row)

    with open(cumulativeCsvName, 'rb') as cumulativeCsvFile:
        cumulativeFileReader = csv.reader(cumulativeCsvFile, delimiter=",", quotechar='"')
        for row in cumulativeFileReader:
            cumulativeSheet.append(row)

    with open(runStreaksCsvName, 'rb') as runStreaksCsvFile:
        runStreaksCsvFileReader = csv.reader(runStreaksCsvFile, delimiter=",", quotechar='"')
        for row in runStreaksCsvFileReader:
            runStreakSheet.append(row)


    # Save the excel file
    excelFileName = fileNameNoExtension + '_FINAL.xlsx'
    wb.save(excelFileName)

    print "Complete. Data output to: %s" % (excelFileName)

else:
    print USAGE